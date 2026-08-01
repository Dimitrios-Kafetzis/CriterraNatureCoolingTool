"""The XLSX workbook: sheet structure, version stamps, and byte-determinism."""

from __future__ import annotations

import zipfile
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from nature_cooling.report import build_xlsx_report
from nature_cooling.report.catalog import FIELD_LABELS, STRINGS
from nature_cooling.report.content import build_content


def _cells(sheet: Worksheet) -> list[tuple[Any, ...]]:
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


def test_workbook_has_the_three_sheets_with_version_stamps(
    render_args: Any, project_name: str
) -> None:
    args = render_args("s01_temperate_street_trees_worked_example")
    result = args["result"]
    workbook = load_workbook(BytesIO(build_xlsx_report(**args)))
    assert workbook.sheetnames == [
        STRINGS["sheet_inputs"],
        STRINGS["sheet_results"],
        STRINGS["sheet_assumptions"],
    ]
    for name in workbook.sheetnames:
        stamp = workbook[name].cell(row=2, column=1).value
        assert stamp is not None
        assert result["methodology_version"] in stamp
        assert result["engine_version"] in stamp
        banner = workbook[name].cell(row=1, column=1).value
        assert project_name in banner
        assert args["label"] in banner


def test_document_timestamps_derive_from_created_at_never_the_clock(
    render_args: Any, created_at: str
) -> None:
    payload = build_xlsx_report(**render_args("s01_temperate_street_trees_worked_example"))
    workbook = load_workbook(BytesIO(payload))
    expected = datetime.fromisoformat(created_at).replace(tzinfo=None)
    assert workbook.properties.created == expected.replace(microsecond=0)
    assert workbook.properties.modified == expected.replace(microsecond=0)
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        stamps = {item.date_time for item in archive.infolist()}
    assert stamps == {(2026, 7, 30, 0, 0, 0)}


def test_zip_member_dates_clamp_to_the_zip_epoch(render_args: Any) -> None:
    args = render_args("s01_temperate_street_trees_worked_example")
    args["created_at"] = "1975-06-01T00:00:00+00:00"
    with zipfile.ZipFile(BytesIO(build_xlsx_report(**args))) as archive:
        stamps = {item.date_time for item in archive.infolist()}
    assert stamps == {(1980, 6, 1, 0, 0, 0)}


def test_inputs_sheet_lists_field_value_and_default_marker(render_args: Any) -> None:
    args = render_args("s05_minimal_input_pocket_park")
    content = build_content(**args)
    sheet = load_workbook(BytesIO(build_xlsx_report(**args)))[STRINGS["sheet_inputs"]]
    rows = _cells(sheet)
    header_index = rows.index(
        (STRINGS["inputs_field"], STRINGS["inputs_value"], STRINGS["inputs_marker"])
    )
    data = rows[header_index + 1 :]
    assert [row[0] for row in data] == [input_row.label for input_row in content.inputs]
    for sheet_row, input_row in zip(data, content.inputs, strict=True):
        assert (sheet_row[1] or "") == input_row.value
        assert (sheet_row[2] or "") == input_row.marker
    assert any(row[2] == STRINGS["not_supplied_marker"] for row in data)
    assert any(row[1] is not None and str(row[1]).endswith("m²") for row in data)


def test_results_sheet_flattens_every_block_row(render_args: Any) -> None:
    args = render_args("s16_urban_forest_short_payback_high_readiness")
    content = build_content(**args)
    sheet = load_workbook(BytesIO(build_xlsx_report(**args)))[STRINGS["sheet_results"]]
    rows = _cells(sheet)
    flat = {(row[0], row[1] or "", row[2]) for row in rows}
    for block in content.blocks:
        for block_row in block.rows:
            assert (block.title, block_row.label, block_row.value) in flat
        for sub_row in block.sub_rows:
            assert (block.title, sub_row.label, sub_row.value) in flat
    assert (STRINGS["recommendation_heading"], "", content.recommendation) in flat
    assert (STRINGS["method_note_heading"], "", content.method_note) in flat
    heat_card = content.cards[0]
    assert (heat_card.title, STRINGS["category"], heat_card.category) in flat
    component_rows = [row for row in rows if row[0] == STRINGS["components_heading"]]
    assert len(component_rows) == len(content.components)


def test_results_sheet_itemises_every_package_component(render_args: Any) -> None:
    """D-038: the workbook carries the component breakdown, not only the headline.

    The workbook has no page budget, so it itemises a package of one as readily
    as a package of five — the PDF's decision to skip the table for a single
    intervention is a layout choice, not a data one.
    """
    args = render_args("s02_tropical_wet_informal_settlement_package")
    content = build_content(**args)
    sheet = load_workbook(BytesIO(build_xlsx_report(**args)))[STRINGS["sheet_results"]]
    rows = _cells(sheet)
    package_rows = [row for row in rows if row[0] == STRINGS["package_heading"]]
    # One row per component, plus the combination-rule note.
    assert len(package_rows) == len(content.package_rows) + 1
    assert package_rows[-1][2] == content.package_note

    by_name = {row[1]: row[2] for row in package_rows[:-1]}
    for component in content.package_rows:
        rendered = by_name[component.name]
        assert component.archetype in rendered
        assert component.evidence in rendered
        assert component.cooling_range in rendered
        # The workbook has no page limit, so it always carries every field.
        assert component.suitability in rendered
        marked = STRINGS["package_representative"] in rendered
        assert marked == component.is_representative


def test_results_sheet_states_the_single_intervention_case_too(render_args: Any) -> None:
    args = render_args("s01_temperate_street_trees_worked_example")
    sheet = load_workbook(BytesIO(build_xlsx_report(**args)))[STRINGS["sheet_results"]]
    values = {row[2] for row in _cells(sheet)}
    assert STRINGS["package_single"] in values


def test_inputs_sheet_labels_the_availability_only_answers(render_args: Any) -> None:
    """D-044.1: the four gating answers are disclosed as feeding no score."""
    args = render_args("s24_productive_landscape_evidence_gap")
    sheet = load_workbook(BytesIO(build_xlsx_report(**args)))[STRINGS["sheet_inputs"]]
    labels = [row[0] for row in _cells(sheet)]
    for field in (
        "includes_railway",
        "existing_woodland",
        "waterfront_type",
        "productive_governance",
    ):
        assert FIELD_LABELS[field] in labels


def test_results_sheet_carries_flags_notes_and_exclusions(render_args: Any) -> None:
    args = render_args("s19_explicit_unknowns_everywhere")
    content = build_content(**args)
    sheet = load_workbook(BytesIO(build_xlsx_report(**args)))[STRINGS["sheet_results"]]
    values = {row[2] for row in _cells(sheet)}
    energy_note = content.blocks[1].note
    assert energy_note in values
    assert content.components_excluded in values

    flagged = render_args("s04_unsuitable_riparian_small_dry_site")
    flagged_content = build_content(**flagged)
    flagged_sheet = load_workbook(BytesIO(build_xlsx_report(**flagged)))[STRINGS["sheet_results"]]
    flagged_values = {row[2] for row in _cells(flagged_sheet)}
    for flag in flagged_content.flags:
        assert flag in flagged_values


def test_assumptions_sheet_itemises_assumptions_and_warnings(render_args: Any) -> None:
    args = render_args("s20_validation_warnings")
    result = args["result"]
    sheet = load_workbook(BytesIO(build_xlsx_report(**args)))[STRINGS["sheet_assumptions"]]
    values = [row[0] for row in _cells(sheet) if row[0] is not None]
    for assumption in result["assumptions_applied"]:
        assert assumption in values
    for warning in result["warnings"]:
        assert warning in values
    assert values.index(STRINGS["assumptions_heading"]) < values.index(STRINGS["warnings_heading"])


def test_assumptions_sheet_states_when_no_default_was_applied(render_args: Any) -> None:
    args = render_args("s01_temperate_street_trees_worked_example")
    args["result"]["assumptions_applied"] = []
    sheet = load_workbook(BytesIO(build_xlsx_report(**args)))[STRINGS["sheet_assumptions"]]
    values = [row[0] for row in _cells(sheet)]
    assert STRINGS["assumptions_none"] in values


def test_same_stored_assessment_produces_byte_identical_workbooks(
    render_args: Any, scenario_names: list[str]
) -> None:
    for name in scenario_names:
        args = render_args(name)
        assert build_xlsx_report(**args) == build_xlsx_report(**args), name
