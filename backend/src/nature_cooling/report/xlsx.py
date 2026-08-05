"""The XLSX workbook: Inputs, Results, Assumptions & Warnings (D-033).

Byte-determinism is a contract: the same stored assessment must produce
byte-identical workbook bytes. Two clock sources have to be pinned for that:

* the document properties (``created`` / ``modified``): openpyxl stamps
  ``modified`` with the wall clock at save time, so the archive's
  ``docProps/core.xml`` is rewritten with both taken from the assessment's
  ``created_at``, keeping clocks confined to the storage layer; and
* the ZIP member timestamps, which :mod:`zipfile` stamps with the wall clock
  on write — the archive is rewritten with every member timestamped from
  ``created_at`` instead (ZIP timestamps cannot precede 1980, so earlier
  dates clamp to the format's epoch).
"""

from __future__ import annotations

import re
import zipfile
from collections.abc import Sequence
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from nature_cooling.report.catalog import STRINGS
from nature_cooling.report.content import (
    ComparisonContent,
    ReportContent,
    ScenarioSource,
    build_comparison_content,
    build_content,
)

_HEADER = Font(bold=True)
_STAMP = Font(italic=True, size=9)


def _deterministic_zip(archive: bytes, stamp: datetime) -> bytes:
    """Rewrite the archive with every timestamp taken from the assessment."""
    date_time = (max(stamp.year, 1980), stamp.month, stamp.day, 0, 0, 0)
    modified = (
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">'
        f"{stamp.replace(microsecond=0).isoformat()}Z</dcterms:modified>"
    ).encode()
    output = BytesIO()
    with (
        zipfile.ZipFile(BytesIO(archive)) as source,
        zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "docProps/core.xml":
                # openpyxl stamps dcterms:modified with the clock at save time.
                payload = re.sub(
                    rb"<dcterms:modified[^>]*>[^<]*</dcterms:modified>", modified, payload
                )
            info = zipfile.ZipInfo(item.filename, date_time=date_time)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = (0o600 << 16) | (item.external_attr & 0xFFFF)
            target.writestr(info, payload)
    return output.getvalue()


def _stamp_sheet(sheet: Worksheet, content: ReportContent) -> int:
    """The identity-and-version banner every sheet carries; returns the next row."""
    title = (
        f"{STRINGS['app_title']} — {STRINGS['app_subtitle']} · "
        f"{content.project_name} · {content.label}"
    )
    sheet.cell(row=1, column=1, value=title).font = _HEADER
    sheet.cell(row=2, column=1, value=content.version_line).font = _STAMP
    return 4


def _write_header(sheet: Worksheet, row: int, headers: tuple[str, ...]) -> int:
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=column, value=header).font = _HEADER
    return row + 1


def _autosize(sheet: Worksheet, widths: tuple[int, ...]) -> None:
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def _inputs_sheet(sheet: Worksheet, content: ReportContent) -> None:
    row = _stamp_sheet(sheet, content)
    sheet.cell(row=row, column=1, value=STRINGS["inputs_note"]).font = _STAMP
    row += 2
    row = _write_header(
        sheet, row, (STRINGS["inputs_field"], STRINGS["inputs_value"], STRINGS["inputs_marker"])
    )
    for input_row in content.inputs:
        sheet.cell(row=row, column=1, value=input_row.label)
        sheet.cell(row=row, column=2, value=input_row.value)
        sheet.cell(row=row, column=3, value=input_row.marker)
        row += 1
    _autosize(sheet, (42, 34, 44))


def _results_sheet(sheet: Worksheet, content: ReportContent) -> None:
    row = _stamp_sheet(sheet, content)
    row = _write_header(
        sheet, row, (STRINGS["results_block"], STRINGS["results_item"], STRINGS["results_value"])
    )

    def write(block: str, item: str, value: str) -> None:
        nonlocal row
        sheet.cell(row=row, column=1, value=block)
        sheet.cell(row=row, column=2, value=item)
        sheet.cell(row=row, column=3, value=value)
        row += 1

    for card in content.cards:
        write(card.title, STRINGS["score"], f"{card.score} {card.scale_note}")
        write(card.title, STRINGS["category"], card.category)
        write(card.title, STRINGS["confidence_heading"], card.confidence)
        for extra in card.extra_rows:
            write(card.title, extra.label, extra.value)
    for flag in content.flags:
        write(STRINGS["flags_heading"], "", flag)
    write(STRINGS["recommendation_heading"], "", content.recommendation)
    for confidence_row in content.confidence_rows:
        write(STRINGS["confidence_heading"], confidence_row.label, confidence_row.value)
    for block in content.blocks:
        if block.confidence is not None:
            write(block.title, STRINGS["confidence_heading"], block.confidence)
        for block_row in block.rows:
            write(block.title, block_row.label, block_row.value)
        for sub_row in block.sub_rows:
            write(block.title, sub_row.label, sub_row.value)
        if block.note is not None:
            write(block.title, "", block.note)
    for component in content.package_rows:
        marker = f" · {STRINGS['package_representative']}" if component.is_representative else ""
        write(
            STRINGS["package_heading"],
            component.name,
            f"{STRINGS['package_archetype']} {component.archetype} · "
            f"{STRINGS['package_evidence']} {component.evidence} · "
            f"{STRINGS['package_range']} {component.cooling_range} · "
            f"{STRINGS['package_suitability']} {component.suitability}{marker}",
        )
    write(STRINGS["package_heading"], "", content.package_note)
    for name, score, nominal, applied in content.components:
        write(
            STRINGS["components_heading"],
            name,
            f"{STRINGS['score']} {score} · {STRINGS['component_nominal_weight']} {nominal} · "
            f"{STRINGS['component_applied_weight']} {applied}",
        )
    if content.components_excluded is not None:
        write(STRINGS["components_heading"], "", content.components_excluded)
    write(STRINGS["method_note_heading"], "", content.method_note)
    _autosize(sheet, (34, 40, 72))


def _assumptions_sheet(sheet: Worksheet, content: ReportContent) -> None:
    row = _stamp_sheet(sheet, content)
    sheet.cell(row=row, column=1, value=STRINGS["assumptions_heading"]).font = _HEADER
    row += 1
    if content.assumptions:
        sheet.cell(row=row, column=1, value=STRINGS["assumptions_intro"]).font = _STAMP
        row += 1
        for assumption in content.assumptions:
            sheet.cell(row=row, column=1, value=assumption)
            row += 1
    else:
        sheet.cell(row=row, column=1, value=STRINGS["assumptions_none"])
        row += 1
    row += 1
    sheet.cell(row=row, column=1, value=STRINGS["warnings_heading"]).font = _HEADER
    row += 1
    for warning in content.warnings:
        sheet.cell(row=row, column=1, value=warning)
        row += 1
    _autosize(sheet, (110,))


# --- Comparison workbook (v2.4) ----------------------------------------------


def _comparison_stamp(sheet: Worksheet, content: ComparisonContent) -> int:
    """The comparison banner every sheet carries; returns the next row."""
    title = (
        f"{STRINGS['app_title']} — {STRINGS['app_subtitle']} · "
        f"{content.project_name} · {STRINGS['comparison_title']}"
    )
    sheet.cell(row=1, column=1, value=title).font = _HEADER
    stamp = f"{STRINGS['created']} {content.created_date}"
    sheet.cell(row=2, column=1, value=stamp).font = _STAMP
    return 4


def _comparison_overview_sheet(sheet: Worksheet, content: ComparisonContent) -> None:
    row = _comparison_stamp(sheet, content)
    row = _write_header(
        sheet,
        row,
        (
            STRINGS["comparison_scenario_column"],
            STRINGS["typology"],
            STRINGS["comparison_scenario_scale"],
            STRINGS["comparison_scenario_created"],
            STRINGS["comparison_versions_column"],
        ),
    )
    for scenario in content.scenarios:
        sheet.cell(row=row, column=1, value=scenario.label)
        sheet.cell(row=row, column=2, value=scenario.typology)
        sheet.cell(row=row, column=3, value=scenario.scale)
        sheet.cell(row=row, column=4, value=scenario.created_date)
        sheet.cell(row=row, column=5, value=scenario.version_line)
        row += 1
    row += 1
    if content.cross_scale_note is not None:
        sheet.cell(row=row, column=1, value=content.cross_scale_note).font = _HEADER
        row += 2
    if content.methodology_note is not None:
        sheet.cell(row=row, column=1, value=content.methodology_note).font = _STAMP
        row += 2
    if content.narrative:
        sheet.cell(row=row, column=1, value=STRINGS["comparison_narrative_heading"]).font = _HEADER
        sheet.cell(row=row, column=2, value=content.narrative)
        row += 2
    labels = tuple(scenario.label for scenario in content.scenarios)
    row = _write_header(sheet, row, (STRINGS["comparison_criterion"], *labels))
    for comparison_row in content.rows:
        sheet.cell(row=row, column=1, value=comparison_row.label)
        for column, (value, best) in enumerate(
            zip(comparison_row.values, comparison_row.best, strict=True), start=2
        ):
            shown = f"{STRINGS['comparison_best_marker']}{value}" if best else value
            cell = sheet.cell(row=row, column=column, value=shown)
            if best:
                cell.font = _HEADER
        row += 1
    row += 1
    if content.like_for_like:
        sheet.cell(row=row, column=1, value=STRINGS["comparison_best_note"]).font = _STAMP
        row += 1
    sheet.cell(row=row, column=1, value=STRINGS["comparison_identical_note"]).font = _STAMP
    # Five columns cover both tables on this sheet: the scenario overview uses
    # all five, the metric table uses one label column plus 2–4 scenarios.
    _autosize(sheet, (34, 32, 30, 30, 40))


def _comparison_site_sheet(sheet: Worksheet, content: ComparisonContent) -> None:
    row = _comparison_stamp(sheet, content)
    sheet.cell(row=row, column=1, value=STRINGS["comparison_site_note"]).font = _STAMP
    row += 2
    row = _write_header(
        sheet, row, (STRINGS["inputs_field"], STRINGS["inputs_value"], STRINGS["inputs_marker"])
    )
    for input_row in content.site_rows:
        sheet.cell(row=row, column=1, value=input_row.label)
        sheet.cell(row=row, column=2, value=input_row.value)
        sheet.cell(row=row, column=3, value=input_row.marker)
        row += 1
    if content.site_differences:
        row += 1
        sheet.cell(
            row=row, column=1, value=STRINGS["comparison_site_differs_heading"]
        ).font = _HEADER
        row += 1
        sheet.cell(row=row, column=1, value=STRINGS["comparison_site_differs_note"]).font = _STAMP
        row += 1
        labels = tuple(scenario.label for scenario in content.scenarios)
        row = _write_header(sheet, row, (STRINGS["inputs_field"], *labels))
        for difference in content.site_differences:
            sheet.cell(row=row, column=1, value=difference.label)
            for column, value in enumerate(difference.values, start=2):
                sheet.cell(row=row, column=column, value=value)
            row += 1
    _autosize(sheet, (42, 34, 44))


def _comparison_detail_sheet(sheet: Worksheet, content: ComparisonContent) -> None:
    row = _comparison_stamp(sheet, content)
    row = _write_header(
        sheet,
        row,
        (
            STRINGS["comparison_scenario_column"],
            STRINGS["comparison_kind"],
            STRINGS["comparison_text_column"],
        ),
    )

    def write(scenario_label: str, kind: str, text: str) -> None:
        nonlocal row
        sheet.cell(row=row, column=1, value=scenario_label)
        sheet.cell(row=row, column=2, value=kind)
        sheet.cell(row=row, column=3, value=text)
        row += 1

    for scenario in content.scenarios:
        if (
            not scenario.flags
            and not scenario.assumptions
            and not scenario.warnings
            and scenario.method_note is None
        ):
            write(scenario.label, "", STRINGS["comparison_no_scenario_detail"])
            continue
        for flag in scenario.flags:
            write(scenario.label, STRINGS["comparison_kind_flag"], flag)
        for assumption in scenario.assumptions:
            write(scenario.label, STRINGS["comparison_kind_assumption"], assumption)
        for warning in scenario.warnings:
            write(scenario.label, STRINGS["comparison_kind_warning"], warning)
        if scenario.method_note is not None:
            write(scenario.label, STRINGS["method_note_heading"], scenario.method_note)
    if content.shared_method_note is not None:
        row += 1
        write("", STRINGS["method_note_heading"], content.shared_method_note)
    _autosize(sheet, (24, 26, 110))


def build_comparison_xlsx_report(
    *, project_name: str, scenarios: Sequence[ScenarioSource]
) -> bytes:
    """Render 2–4 stored, evaluated assessments as one comparison workbook."""
    content = build_comparison_content(project_name=project_name, scenarios=scenarios)
    stamp = datetime.fromisoformat(content.created_at).replace(tzinfo=None)

    workbook = Workbook()
    workbook.properties.creator = "Criterra"
    workbook.properties.created = stamp
    workbook.properties.modified = stamp
    workbook.properties.title = (
        f"{STRINGS['app_title']} {STRINGS['comparison_title']} — {project_name}"
    )

    overview = workbook.active
    assert overview is not None  # a fresh workbook always has an active sheet
    overview.title = STRINGS["comparison_sheet_overview"]
    _comparison_overview_sheet(overview, content)
    _comparison_site_sheet(workbook.create_sheet(STRINGS["comparison_sheet_site"]), content)
    _comparison_detail_sheet(workbook.create_sheet(STRINGS["comparison_sheet_scenarios"]), content)

    buffer = BytesIO()
    workbook.save(buffer)
    return _deterministic_zip(buffer.getvalue(), stamp)


def build_xlsx_report(
    *,
    project_name: str,
    label: str,
    created_at: str,
    inp: dict[str, Any],
    result: dict[str, Any],
    autofilled: dict[str, str] | None = None,
) -> bytes:
    """Render one stored, evaluated assessment as workbook bytes."""
    content = build_content(
        project_name=project_name,
        label=label,
        created_at=created_at,
        inp=inp,
        result=result,
        autofilled=autofilled,
    )
    stamp = datetime.fromisoformat(created_at).replace(tzinfo=None)

    workbook = Workbook()
    workbook.properties.creator = "Criterra"
    workbook.properties.created = stamp
    workbook.properties.modified = stamp
    workbook.properties.title = (
        f"{STRINGS['app_title']} {STRINGS['report_title']} — {project_name} · {label}"
    )

    inputs = workbook.active
    assert inputs is not None  # a fresh workbook always has an active sheet
    inputs.title = STRINGS["sheet_inputs"]
    _inputs_sheet(inputs, content)
    _results_sheet(workbook.create_sheet(STRINGS["sheet_results"]), content)
    _assumptions_sheet(workbook.create_sheet(STRINGS["sheet_assumptions"]), content)

    buffer = BytesIO()
    workbook.save(buffer)
    return _deterministic_zip(buffer.getvalue(), stamp)
