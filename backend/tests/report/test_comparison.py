"""The v2.4 comparison report: content model, PDF, and workbook.

The verdict-style contract under test: best-per-criterion marking plus a
factual narrative — no overall winner, no ranking, no recommendation — with
marking withheld entirely where the scenarios are not like for like or where
a status competes with a figure.
"""

from __future__ import annotations

import zipfile
from datetime import datetime
from io import BytesIO
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pypdf import PdfReader

from nature_cooling.report import (
    build_comparison_pdf_report,
    build_comparison_xlsx_report,
)
from nature_cooling.report.catalog import FIELD_LABELS, STRINGS
from nature_cooling.report.content import build_comparison_content

S01 = "s01_temperate_street_trees_worked_example"
S02 = "s02_tropical_wet_informal_settlement_package"
S03 = "s03_arid_urban_forest_ceiling_clip"
S16 = "s16_urban_forest_short_payback_high_readiness"


def _pdf_text(payload: bytes) -> str:
    reader = PdfReader(BytesIO(payload))
    return " ".join(" ".join(page.extract_text() for page in reader.pages).split())


def _cells(sheet: Worksheet) -> list[tuple[Any, ...]]:
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


def _row(content: Any, label: str) -> Any:
    return next(row for row in content.rows if row.label == label)


# --- Content model -----------------------------------------------------------


def test_a_comparison_takes_two_to_four_scenarios(comparison_source: Any) -> None:
    one = [comparison_source(S01, "Only")]
    with pytest.raises(ValueError, match="2 to 4"):
        build_comparison_content(project_name="P", scenarios=one)
    five = [comparison_source(S01, f"Option {index}") for index in range(5)]
    with pytest.raises(ValueError, match="2 to 4"):
        build_comparison_content(project_name="P", scenarios=five)


def test_best_per_criterion_marking_and_the_factual_narrative(
    comparison_source: Any,
) -> None:
    base = comparison_source(S01, "Street trees")
    other = comparison_source(S01, "More trees")
    other.result["opportunity"]["score"] = 80.0
    other.result["cooling"]["potential_score"] = 90.0
    content = build_comparison_content(project_name="Riverside", scenarios=[base, other])

    assert content.like_for_like
    assert content.cross_scale_note is None
    assert content.methodology_note is None
    opportunity = _row(content, STRINGS["opportunity"])
    assert opportunity.best == (False, True)
    assert opportunity.differs
    cooling = _row(content, STRINGS["comparison_cooling_potential"])
    assert cooling.best == (False, True)
    # The narrative states facts the table marks, naming the criterion as the
    # table labels it — and never says "choose".
    assert "More trees has the highest NbS Cooling Opportunity Score (80)" in content.narrative
    assert "More trees has the highest Cooling Potential Score (90)" in content.narrative
    assert "choose" not in content.narrative.lower()
    # Identical criteria are neither marked nor narrated.
    delta_t = _row(content, STRINGS["cooling_delta_t"])
    assert delta_t.best == (False, False)
    assert not delta_t.differs


def test_the_heat_priority_index_is_never_marked_best(comparison_source: Any) -> None:
    """A hotter site is not a better scenario: the index has no direction."""
    base = comparison_source(S01, "A")
    other = comparison_source(S01, "B")
    other.result["heat_priority"]["score"] = 95.0
    content = build_comparison_content(project_name="P", scenarios=[base, other])
    row = _row(content, STRINGS["heat_priority"])
    assert row.differs
    assert row.best == (False, False)
    assert STRINGS["heat_priority"] not in content.narrative


def test_lower_is_better_for_payback(comparison_source: Any) -> None:
    base = comparison_source(S16, "Forest")
    other = comparison_source(S16, "Forest, phased")
    other.result["costs"]["payback_years_central"] = (
        base.result["costs"]["payback_years_central"] + 2.0
    )
    content = build_comparison_content(project_name="P", scenarios=[base, other])
    row = _row(content, STRINGS["costs_payback"])
    assert row.best == (True, False)
    assert f"Forest has the lowest {STRINGS['costs_payback']}" in content.narrative


def test_ties_are_marked_together_and_named_together(comparison_source: Any) -> None:
    first = comparison_source(S01, "A")
    second = comparison_source(S01, "B")
    third = comparison_source(S01, "C")
    first.result["opportunity"]["score"] = 80.0
    second.result["opportunity"]["score"] = 80.0
    content = build_comparison_content(project_name="P", scenarios=[first, second, third])
    row = _row(content, STRINGS["opportunity"])
    assert row.best == (True, True, False)
    assert f"A and B share the highest {STRINGS['opportunity']} (80)" in content.narrative


def test_identical_scenarios_state_the_empty_narrative(comparison_source: Any) -> None:
    content = build_comparison_content(
        project_name="P",
        scenarios=[comparison_source(S01, "A"), comparison_source(S01, "B")],
    )
    assert content.narrative == STRINGS["comparison_narrative_empty"]
    assert all(not row.differs for row in content.rows)
    assert all(row.best == (False, False) for row in content.rows)


def test_a_status_never_competes_with_a_figure(comparison_source: Any) -> None:
    """ "Not estimated" cannot lose to a number: the criterion goes unmarked."""
    base = comparison_source(S16, "A")
    other = comparison_source(S16, "B")
    other.result["energy"]["status"] = "missing_energy_demand"
    other.result["energy"]["status_message"] = "Not estimated — demand was not provided."
    content = build_comparison_content(project_name="P", scenarios=[base, other])
    row = _row(content, STRINGS["energy_savings"])
    assert row.differs
    assert row.best == (False, False)
    # The engine's own words render verbatim.
    assert row.values[1] == "Not estimated — demand was not provided."
    assert STRINGS["energy_savings"] not in content.narrative


def test_overall_confidence_is_ordered_and_unknown_levels_never_marked(
    comparison_source: Any,
) -> None:
    base = comparison_source(S16, "A")
    other = comparison_source(S16, "B")
    other.result["confidence"]["overall"] = "low"
    content = build_comparison_content(project_name="P", scenarios=[base, other])
    row = _row(content, STRINGS["comparison_confidence"])
    assert row.best == (True, False)

    # A level this build does not know cannot honestly be ranked.
    odd = comparison_source(S16, "B")
    odd.result["confidence"]["overall"] = "experimental"
    content = build_comparison_content(project_name="P", scenarios=[base, odd])
    row = _row(content, STRINGS["comparison_confidence"])
    assert row.differs
    assert row.best == (False, False)


def test_cross_scale_scenarios_are_flagged_and_never_marked(
    comparison_source: Any,
) -> None:
    """The brief's honesty rule: never silently tabulated as like for like."""
    site = comparison_source(S02, "Settlement package")
    district = comparison_source(S03, "Urban forest")
    content = build_comparison_content(project_name="P", scenarios=[site, district])
    assert not content.like_for_like
    assert content.cross_scale_note is not None
    assert "Settlement package: Site" in content.cross_scale_note
    assert "Urban forest: District" in content.cross_scale_note
    assert content.narrative == ""
    assert all(row.best == (False, False) for row in content.rows)


def test_differing_methodology_versions_are_disclosed_per_scenario(
    comparison_source: Any,
) -> None:
    base = comparison_source(S01, "A")
    older = comparison_source(S01, "B")
    older.result["methodology_version"] = "2026.07.30"
    older.result["method_note"] = "An earlier method note."
    content = build_comparison_content(project_name="P", scenarios=[base, older])
    assert content.methodology_note is not None
    assert "A: " + str(base.result["methodology_version"]) in content.methodology_note
    assert "B: 2026.07.30" in content.methodology_note
    # Method notes differ, so each renders with its own scenario.
    assert content.shared_method_note is None
    assert content.scenarios[0].method_note == str(base.result["method_note"])
    assert content.scenarios[1].method_note == "An earlier method note."
    # Either way, every scenario's overview names what it was evaluated under.
    for scenario, source in zip(content.scenarios, (base, older), strict=True):
        assert str(source.result["methodology_version"]) in scenario.version_line
        assert str(source.result["engine_version"]) in scenario.version_line


def test_a_shared_method_note_renders_once(comparison_source: Any) -> None:
    content = build_comparison_content(
        project_name="P",
        scenarios=[comparison_source(S01, "A"), comparison_source(S01, "B")],
    )
    assert content.shared_method_note == str(comparison_source(S01, "A").result["method_note"])
    assert all(scenario.method_note is None for scenario in content.scenarios)


def test_site_context_is_printed_once_and_differences_are_disclosed(
    comparison_source: Any,
) -> None:
    base = comparison_source(S01, "A")
    other = comparison_source(S01, "B")
    other.inp["existing_tree_canopy_percent"] = 30.0
    other.inp["nbs_type"] = ["pocket_park"]
    content = build_comparison_content(project_name="P", scenarios=[base, other])

    shared_fields = {row.field for row in content.site_rows}
    assert "site_area_m2" in shared_fields
    # The like-for-like axis is disclosed in the overview, not the context.
    assert "assessment_scale" not in shared_fields
    # Intervention and cost answers are not site context, differing or not.
    differing = {row.label for row in content.site_differences}
    assert FIELD_LABELS["nbs_type"] not in differing
    assert FIELD_LABELS["existing_tree_canopy_percent"] in differing
    assert FIELD_LABELS["existing_tree_canopy_percent"] not in {
        row.label for row in content.site_rows
    }


def test_shared_values_carry_their_provenance_or_say_it_differs(
    comparison_source: Any,
) -> None:
    both_marked = build_comparison_content(
        project_name="P",
        scenarios=[
            comparison_source(S01, "A", autofilled={"climate_zone": "beck2023"}),
            comparison_source(S01, "B", autofilled={"climate_zone": "beck2023"}),
        ],
    )
    climate = next(row for row in both_marked.site_rows if row.field == "climate_zone")
    assert "Beck et al. 2023" in climate.marker

    one_marked = build_comparison_content(
        project_name="P",
        scenarios=[
            comparison_source(S01, "A", autofilled={"climate_zone": "beck2023"}),
            comparison_source(S01, "B"),
        ],
    )
    climate = next(row for row in one_marked.site_rows if row.field == "climate_zone")
    assert climate.marker == STRINGS["provenance_differs"]


def test_the_comparison_clock_is_the_newest_scenario(comparison_source: Any) -> None:
    early = comparison_source(S01, "A", created_at="2026-08-01T10:00:00+00:00")
    late = comparison_source(S01, "B", created_at="2026-08-03T09:30:00+00:00")
    content = build_comparison_content(project_name="P", scenarios=[late, early])
    assert content.created_at == "2026-08-03T09:30:00+00:00"
    assert content.created_date == "2026-08-03"
    assert content.scenarios[1].created_date == "2026-08-01"


# --- PDF ---------------------------------------------------------------------


def test_pdf_carries_overview_table_site_context_and_scenario_detail(
    comparison_source: Any,
) -> None:
    base = comparison_source(S01, "Street trees")
    other = comparison_source(S01, "More trees")
    other.result["opportunity"]["score"] = 80.0
    content = build_comparison_content(project_name="Riverside", scenarios=[base, other])
    text = _pdf_text(build_comparison_pdf_report(project_name="Riverside", scenarios=[base, other]))
    assert "Riverside" in text
    assert STRINGS["comparison_scenarios_heading"].upper() in text
    assert STRINGS["comparison_table_heading"].upper() in text
    assert content.narrative[:60] in text
    # Both scenario labels head the table and the overview.
    assert "Street trees" in text and "More trees" in text
    # The best cell carries the marker; the notes explain it.
    assert f"{STRINGS['comparison_best_marker']}80" in text
    assert STRINGS["comparison_best_note"][:40] in text
    assert STRINGS["comparison_identical_note"][:40] in text
    assert STRINGS["comparison_site_heading"].upper() in text
    for scenario in content.scenarios:
        assert STRINGS["comparison_scenario_detail"].format(label=scenario.label).upper() in text
    assert content.shared_method_note is not None
    assert content.shared_method_note[:50] in text
    assert STRINGS["copyright_line"] in text


def test_pdf_flags_cross_scale_comparisons_and_marks_nothing(
    comparison_source: Any,
) -> None:
    scenarios = [
        comparison_source(S02, "Settlement package"),
        comparison_source(S03, "Urban forest"),
    ]
    content = build_comparison_content(project_name="P", scenarios=scenarios)
    text = _pdf_text(build_comparison_pdf_report(project_name="P", scenarios=scenarios))
    assert content.cross_scale_note is not None
    assert content.cross_scale_note[:60] in text
    assert STRINGS["comparison_best_marker"].strip() + " " not in text
    assert STRINGS["comparison_narrative_heading"].upper() not in text
    assert STRINGS["comparison_best_note"][:40] not in text


def test_pdf_discloses_differing_methodology_versions_and_site_answers(
    comparison_source: Any,
) -> None:
    base = comparison_source(S01, "A")
    other = comparison_source(S01, "B")
    other.result["methodology_version"] = "2026.07.30"
    other.result["method_note"] = "An earlier method note."
    other.inp["existing_tree_canopy_percent"] = 30.0
    content = build_comparison_content(project_name="P", scenarios=[base, other])
    text = _pdf_text(build_comparison_pdf_report(project_name="P", scenarios=[base, other]))
    assert content.methodology_note is not None
    assert content.methodology_note[:60] in text
    assert STRINGS["comparison_site_differs_heading"].upper() in text
    assert FIELD_LABELS["existing_tree_canopy_percent"] in text
    # Method notes differ → each scenario's own note renders in its section.
    assert "An earlier method note." in text


def test_pdf_states_when_a_scenario_has_nothing_to_itemise(
    comparison_source: Any,
) -> None:
    scenarios = [comparison_source(S01, "A"), comparison_source(S01, "B")]
    for scenario in scenarios:
        scenario.result["assumptions_applied"] = []
        scenario.result["warnings"] = []
    text = _pdf_text(build_comparison_pdf_report(project_name="P", scenarios=scenarios))
    assert STRINGS["comparison_no_scenario_detail"][:50] in text


def test_pdf_itemises_flags_assumptions_and_warnings_per_scenario(
    comparison_source: Any,
) -> None:
    base = comparison_source(S01, "A")
    other = comparison_source(S01, "B")
    other.result["suitability"]["flags"] = [{"code": "x", "message": "Not suitable here."}]
    other.result["warnings"] = ["A recorded warning."]
    text = _pdf_text(build_comparison_pdf_report(project_name="P", scenarios=[base, other]))
    assert "Not suitable here." in text
    assert "A recorded warning." in text
    assert STRINGS["assumptions_intro"][:40] in text


def test_pdf_grows_beyond_one_page_rather_than_truncating(
    comparison_source: Any,
) -> None:
    scenarios = [
        comparison_source(S01, "Street trees"),
        comparison_source(S02, "Settlement package"),
        comparison_source(S03, "Urban forest"),
        comparison_source(S16, "Payback forest"),
    ]
    payload = build_comparison_pdf_report(project_name="P", scenarios=scenarios)
    assert len(PdfReader(BytesIO(payload)).pages) > 1
    text = _pdf_text(payload)
    # Nothing is truncated: every scenario's assumptions all render.
    for scenario in scenarios:
        for assumption in scenario.result["assumptions_applied"]:
            assert assumption[:50] in text


def test_pdf_metadata_derives_from_the_newest_scenario_never_the_clock(
    comparison_source: Any,
) -> None:
    scenarios = [
        comparison_source(S01, "A", created_at="2026-08-01T10:00:00+00:00"),
        comparison_source(S01, "B", created_at="2026-08-03T09:30:00+00:00"),
    ]
    reader = PdfReader(
        BytesIO(build_comparison_pdf_report(project_name="Riverside", scenarios=scenarios))
    )
    metadata = reader.metadata
    assert metadata is not None
    assert metadata.get("/CreationDate") == "D:20260803093000Z"
    assert metadata.get("/Author") == "Criterra"
    assert STRINGS["comparison_title"] in str(metadata.get("/Title"))
    assert "Riverside" in str(metadata.get("/Title"))


def test_same_stored_scenarios_produce_byte_identical_pdfs(
    comparison_source: Any,
) -> None:
    for names in ([S01, S16], [S02, S03]):
        scenarios = [comparison_source(name, f"Option {index}") for index, name in enumerate(names)]
        first = build_comparison_pdf_report(project_name="P", scenarios=scenarios)
        second = build_comparison_pdf_report(project_name="P", scenarios=scenarios)
        assert first == second, names


# --- Workbook ----------------------------------------------------------------


def test_workbook_has_the_three_comparison_sheets_with_banners(
    comparison_source: Any,
) -> None:
    scenarios = [comparison_source(S01, "A"), comparison_source(S01, "B")]
    workbook = load_workbook(
        BytesIO(build_comparison_xlsx_report(project_name="Riverside", scenarios=scenarios))
    )
    assert workbook.sheetnames == [
        STRINGS["comparison_sheet_overview"],
        STRINGS["comparison_sheet_site"],
        STRINGS["comparison_sheet_scenarios"],
    ]
    for name in workbook.sheetnames:
        banner = workbook[name].cell(row=1, column=1).value
        assert "Riverside" in banner
        assert STRINGS["comparison_title"] in banner


def test_workbook_table_keeps_the_callers_column_order_and_marks_best(
    comparison_source: Any,
) -> None:
    base = comparison_source(S01, "Street trees")
    other = comparison_source(S01, "More trees")
    other.result["opportunity"]["score"] = 80.0
    payload = build_comparison_xlsx_report(project_name="P", scenarios=[other, base])
    sheet = load_workbook(BytesIO(payload))[STRINGS["comparison_sheet_overview"]]
    rows = _cells(sheet)
    header = (STRINGS["comparison_criterion"], "More trees", "Street trees")
    assert header in {row[:3] for row in rows}
    opportunity = next(row for row in rows if row[0] == STRINGS["opportunity"])
    assert str(opportunity[1]).startswith(STRINGS["comparison_best_marker"])
    assert not str(opportunity[2]).startswith(STRINGS["comparison_best_marker"])
    flat = [row[0] for row in rows]
    assert STRINGS["comparison_best_note"] in flat
    assert STRINGS["comparison_identical_note"] in flat
    narrative = next(row for row in rows if row[0] == STRINGS["comparison_narrative_heading"])
    assert "More trees has the highest" in str(narrative[1])
    # Column order is the caller's: swapping the call swaps the columns.
    swapped = build_comparison_xlsx_report(project_name="P", scenarios=[base, other])
    assert payload != swapped


def test_workbook_overview_names_scale_and_versions_per_scenario(
    comparison_source: Any,
) -> None:
    scenarios = [
        comparison_source(S02, "Settlement package"),
        comparison_source(S03, "Urban forest"),
    ]
    content = build_comparison_content(project_name="P", scenarios=scenarios)
    sheet = load_workbook(
        BytesIO(build_comparison_xlsx_report(project_name="P", scenarios=scenarios))
    )[STRINGS["comparison_sheet_overview"]]
    rows = _cells(sheet)
    for scenario in content.scenarios:
        entry = next(row for row in rows if row[0] == scenario.label)
        assert entry[1] == scenario.typology
        assert entry[2] == scenario.scale
        assert entry[3] == scenario.created_date
        assert entry[4] == scenario.version_line
    # Cross-scale: the flag row is present, the best-note legend is not.
    flat = [row[0] for row in rows]
    assert content.cross_scale_note in flat
    assert STRINGS["comparison_best_note"] not in flat


def test_workbook_site_sheet_prints_shared_context_once_and_differences(
    comparison_source: Any,
) -> None:
    base = comparison_source(S01, "A")
    other = comparison_source(S01, "B")
    other.inp["existing_tree_canopy_percent"] = 30.0
    other.result["methodology_version"] = "2026.07.30"
    content = build_comparison_content(project_name="P", scenarios=[base, other])
    workbook = load_workbook(
        BytesIO(build_comparison_xlsx_report(project_name="P", scenarios=[base, other]))
    )
    site = _cells(workbook[STRINGS["comparison_sheet_site"]])
    flat = {(row[0], row[1] or "", row[2] or "") for row in site}
    for input_row in content.site_rows:
        assert (input_row.label, input_row.value, input_row.marker) in flat
    difference = next(row for row in site if row[0] == FIELD_LABELS["existing_tree_canopy_percent"])
    assert difference[1] != difference[2]
    # The methodology disclosure travels to the overview sheet.
    overview = [row[0] for row in _cells(workbook[STRINGS["comparison_sheet_overview"]])]
    assert content.methodology_note in overview


def test_workbook_detail_sheet_itemises_per_scenario_and_states_emptiness(
    comparison_source: Any,
) -> None:
    base = comparison_source(S01, "A")
    base.result["assumptions_applied"] = []
    other = comparison_source(S01, "B")
    other.result["suitability"]["flags"] = [{"code": "x", "message": "Not suitable here."}]
    other.result["warnings"] = ["A recorded warning."]
    sheet = load_workbook(
        BytesIO(build_comparison_xlsx_report(project_name="P", scenarios=[base, other]))
    )[STRINGS["comparison_sheet_scenarios"]]
    rows = _cells(sheet)
    # An empty kind cell reads back as None: the row states the emptiness.
    assert ("A", None, STRINGS["comparison_no_scenario_detail"]) in rows
    assert ("B", STRINGS["comparison_kind_flag"], "Not suitable here.") in rows
    assert ("B", STRINGS["comparison_kind_warning"], "A recorded warning.") in rows
    kinds = {row[1] for row in rows if row[0] == "B"}
    assert STRINGS["comparison_kind_assumption"] in kinds
    # The method note is shared, so it renders once, unattributed.
    shared = next(row for row in rows if row[1] == STRINGS["method_note_heading"])
    assert shared[0] is None or shared[0] == ""


def test_workbook_detail_sheet_attributes_differing_method_notes(
    comparison_source: Any,
) -> None:
    base = comparison_source(S01, "A")
    other = comparison_source(S01, "B")
    other.result["method_note"] = "An earlier method note."
    sheet = load_workbook(
        BytesIO(build_comparison_xlsx_report(project_name="P", scenarios=[base, other]))
    )[STRINGS["comparison_sheet_scenarios"]]
    rows = _cells(sheet)
    assert ("B", STRINGS["method_note_heading"], "An earlier method note.") in rows
    assert ("A", STRINGS["method_note_heading"], str(base.result["method_note"])) in rows


def test_workbook_timestamps_derive_from_the_newest_scenario(
    comparison_source: Any,
) -> None:
    scenarios = [
        comparison_source(S01, "A", created_at="2026-08-01T10:00:00+00:00"),
        comparison_source(S01, "B", created_at="2026-08-03T09:30:00+00:00"),
    ]
    payload = build_comparison_xlsx_report(project_name="P", scenarios=scenarios)
    workbook = load_workbook(BytesIO(payload))
    expected = datetime.fromisoformat("2026-08-03T09:30:00+00:00").replace(tzinfo=None)
    assert workbook.properties.created == expected
    assert workbook.properties.modified == expected
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        stamps = {item.date_time for item in archive.infolist()}
    assert stamps == {(2026, 8, 3, 0, 0, 0)}


def test_same_stored_scenarios_produce_byte_identical_workbooks(
    comparison_source: Any,
) -> None:
    for names in ([S01, S16], [S02, S03]):
        scenarios = [comparison_source(name, f"Option {index}") for index, name in enumerate(names)]
        first = build_comparison_xlsx_report(project_name="P", scenarios=scenarios)
        second = build_comparison_xlsx_report(project_name="P", scenarios=scenarios)
        assert first == second, names
